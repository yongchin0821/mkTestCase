#!/usr/bin/python
# -*- coding: utf-8 -*-
# @Time    : 2023/3/6 10:59 AM
# @Author  : Yongchin

from openpyxl import Workbook, load_workbook

if __name__ == '__main__':
    wb = Workbook()
    ws = wb.active
    # ws.cell(i + 3, 15, data)
    # 保存
    wb.save("/Users/xxx/Documents/普通商密/自动化测试跟踪表示例/大屏3.0升级自动化输入表新.xlsx")
